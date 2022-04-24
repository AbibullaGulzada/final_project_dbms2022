from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys #import this to acces simple keyboard keys via selenium
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np

book_df = pd.read_excel (r'D:\Users\Almas\Documents\GitHub\DB_PROJECT_SDU\data_parsing\book-listing(cleaned).xlsx')
#adding new two columns for price and year
book_df["price"] = ""
book_df["date"] = ""

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()




def getPriceAndDate(a_id):
    PATH = "D:\Program Files (x86)\chromedriver.exe" #path for chromedriver executional file so that selenium can run through it.
    c_options = Options()
    chrome_prefs = {}
    c_options.experimental_options["prefs"] = chrome_prefs
    chrome_prefs["profile.default_content_settings"] = {"images": 2}
    chrome_prefs["profile.managed_default_content_settings"] = {"images": 2} #disable image loading fastens speed
    c_options.headless = True #to run it without popup browser
    driver = webdriver.Chrome(PATH, options=c_options)

    driver.get("https://www.amazon.com/s?k=" + str(a_id) + "&i=stripbooks-intl-ship&ref=nb_sb_noss")
    content = driver.page_source
    soup = BeautifulSoup(content, features="html.parser")

    try:
        soup_price = soup.find(name='div', attrs={'data-index':'0'}).find(name='span', attrs={'class':'a-offscreen'}) #finding price
        #<span class="a-offscreen">$7.49</span>
        if soup_price is not None:
            print(soup_price.text)
            price_out = soup_price.text
        else:
            print("Can't find the price")
            price_out = ""
        #finding date below
        soup_date = soup.find(name='div', attrs={'data-index':'0'}).find(name='div', attrs={'class':'a-row'}).find(name='span', attrs={'class':'a-text-normal', 'dir':'auto'})
        #<span class="a-size-base a-color-secondary a-text-normal" dir="auto">Sep 23, 2014</span>
        if soup_date is not None:
            print(soup_date.text)
            date_out = soup_date.text
        else:
            print("Can't find date")
            date_out = ""
    except:
        price_out = ""
        date_out = ""
    if price_out != "" and date_out != "":
        print("succesfully done")
    else:
        print("fail")
    driver.quit()
    return price_out, date_out

#new dataframe for better data.
new_df = pd.DataFrame(columns=['id','url', 'title', 'author', 'genre_id', 'genre_name', 'price', 'date'])

for i in range(7000, 8000):
    print(i+1)
    book_df.at[i, 'price'], book_df.at[i, 'date'] = getPriceAndDate(book_df.values[i, 0])
    new_df = new_df.append(book_df.loc[i], ignore_index=True)
append_df_to_excel(r'D:\Users\Almas\Documents\GitHub\DB_PROJECT_SDU\data_parsing\book-listing(cleaned).xlsx', new_df)


print(new_df)
